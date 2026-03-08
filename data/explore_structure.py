#!/usr/bin/env python3
"""Step 1: Explore the structure of representative estimate xlsx files."""

import openpyxl
import os
import json
import traceback

BASE = "/Users/jameswalton/Desktop/Construction Projects"

# Find all xlsx files with "estimate" or "estiamte" in the name
all_xlsx = []
for root, dirs, files in os.walk(BASE):
    for f in files:
        if f.endswith('.xlsx') and not f.startswith('~$'):
            all_xlsx.append(os.path.join(root, f))

estimate_files = [f for f in all_xlsx if 'estimate' in os.path.basename(f).lower() or 'estiamte' in os.path.basename(f).lower()]

# Also grab Materials Worksheet files - they seem important
materials_files = [f for f in all_xlsx if 'materials worksheet' in os.path.basename(f).lower()]

# And other key types
allowance_files = [f for f in all_xlsx if 'allowance' in os.path.basename(f).lower()]
cost_labor_files = [f for f in all_xlsx if 'cost and labor' in os.path.basename(f).lower() or 'takeoff' in os.path.basename(f).lower()]

print(f"Total xlsx files: {len(all_xlsx)}")
print(f"Estimate files: {len(estimate_files)}")
print(f"Materials Worksheet files: {len(materials_files)}")
print(f"Allowance files: {len(allowance_files)}")
print(f"Cost/Labor/Takeoff files: {len(cost_labor_files)}")

# Pick a diverse sample of ~18 files
sample_files = []

# Pick from different categories
for lst, label in [(estimate_files, "estimate"), (materials_files, "materials"),
                    (allowance_files, "allowance"), (cost_labor_files, "cost_labor")]:
    # Pick up to 5 from each, spread across different projects
    seen_projects = set()
    for f in lst:
        # Extract project folder name
        rel = os.path.relpath(f, BASE)
        parts = rel.split(os.sep)
        project = parts[1] if len(parts) > 1 else parts[0]
        if project not in seen_projects and len([x for x in sample_files if x[1] == label]) < 5:
            sample_files.append((f, label))
            seen_projects.add(project)

print(f"\nSample size: {len(sample_files)}")
print("=" * 80)

results = {}

for filepath, category in sample_files:
    fname = os.path.basename(filepath)
    print(f"\n{'='*80}")
    print(f"FILE: {fname}")
    print(f"CATEGORY: {category}")
    print(f"PATH: {filepath}")

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        file_info = {
            "filename": fname,
            "category": category,
            "sheets": []
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n  SHEET: '{sheet_name}'")

            sheet_info = {
                "name": sheet_name,
                "rows_sample": []
            }

            row_count = 0
            for row in ws.iter_rows(max_row=40, values_only=False):
                row_count += 1
                values = []
                for cell in row:
                    val = cell.value
                    if val is not None:
                        values.append(str(val)[:80])
                    else:
                        values.append("")

                # Only print rows that have some content
                non_empty = [v for v in values if v.strip()]
                if non_empty:
                    print(f"    Row {row_count}: {values[:12]}")
                    sheet_info["rows_sample"].append({
                        "row": row_count,
                        "values": values[:12]
                    })

            file_info["sheets"].append(sheet_info)

        results[fname] = file_info
        wb.close()

    except Exception as e:
        print(f"  ERROR: {e}")
        traceback.print_exc()

# Save structure analysis
with open("/Users/jameswalton/Desktop/MHPEstimate/data/structure_analysis.json", "w") as f:
    json.dump(results, f, indent=2, default=str)

print("\n\nStructure analysis saved.")
