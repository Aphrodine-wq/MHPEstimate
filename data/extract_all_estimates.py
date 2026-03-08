#!/usr/bin/env python3
"""
MHP Construction Estimate Data Extractor
Reads all xlsx files from Construction Projects directory,
extracts pricing data, and outputs consolidated JSON.
"""

import openpyxl
import os
import json
import re
import traceback
from collections import defaultdict
from statistics import mean, median

BASE = "/Users/jameswalton/Desktop/Construction Projects"
OUTPUT_DIR = "/Users/jameswalton/Desktop/MHPEstimate/data"

def find_all_xlsx():
    """Find all xlsx files, excluding temp files."""
    files = []
    for root, dirs, filenames in os.walk(BASE):
        for f in filenames:
            if f.endswith('.xlsx') and not f.startswith('~$'):
                files.append(os.path.join(root, f))
    return files

def classify_file(filepath):
    """Classify file type based on name."""
    name = os.path.basename(filepath).lower()
    if any(k in name for k in ['estimate', 'estiamte']):
        return 'estimate'
    elif 'allowance' in name or 'selection' in name:
        return 'allowance'
    elif 'materials worksheet' in name:
        return 'materials_worksheet'
    elif 'expense' in name or 'payment' in name or 'invoice' in name:
        return 'financial'
    elif 'timeline' in name or 'schedule' in name:
        return 'timeline'
    elif 'worklog' in name or 'daily' in name or 'report' in name:
        return 'operations'
    elif 'contracted work' in name:
        return 'contracted_work'
    elif 'takeoff' in name or 'take-off' in name:
        return 'takeoff'
    else:
        return 'other'

def extract_project_name(filepath):
    """Extract project/client name from path."""
    rel = os.path.relpath(filepath, BASE)
    parts = rel.split(os.sep)
    # Skip OneDrive folder, get project folder
    if len(parts) >= 3:
        folder = parts[1]
        if folder in ('Completed Projects', '__Completed Projects', 'Paused Projects'):
            return parts[2] if len(parts) > 2 else folder
        return folder
    return parts[-1]

def extract_project_type(filepath, project_name):
    """Infer project type from path and filename."""
    combined = (filepath + " " + project_name).lower()
    if any(k in combined for k in ['porch', 'screen porch']):
        return 'porch'
    elif 'deck' in combined:
        return 'deck'
    elif 'kitchen' in combined:
        return 'kitchen_renovation'
    elif any(k in combined for k in ['bathroom', 'shower', 'bath']):
        return 'bathroom_renovation'
    elif 'addition' in combined or 'remodel' in combined:
        return 'addition_remodel'
    elif any(k in combined for k in ['guest house', 'adu', 'cottage']):
        return 'guest_house'
    elif 'new build' in combined or 'new home' in combined:
        return 'new_build'
    elif any(k in combined for k in ['garage', 'carport']):
        return 'garage_carport'
    elif 'retaining wall' in combined:
        return 'retaining_wall'
    elif 'fence' in combined:
        return 'fencing'
    elif 'roof' in combined:
        return 'roofing'
    elif any(k in combined for k in ['paint', 'painting']):
        return 'painting'
    elif any(k in combined for k in ['concrete', 'driveway', 'hardscape']):
        return 'concrete_hardscape'
    elif any(k in combined for k in ['door', 'window']):
        return 'door_window'
    elif 'bonus room' in combined:
        return 'bonus_room'
    elif any(k in combined for k in ['barn', 'pole barn', 'shop']):
        return 'barn_outbuilding'
    elif any(k in combined for k in ['veterinary', 'clinic', 'bank', 'church', 'law']):
        return 'commercial'
    elif 'demo' in combined:
        return 'demolition'
    elif 'gutter' in combined:
        return 'gutter'
    else:
        return 'general'

def is_numeric(val):
    """Check if value is numeric."""
    if val is None:
        return False
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, str):
        val = val.replace('$', '').replace(',', '').strip()
        try:
            float(val)
            return True
        except:
            return False
    return False

def to_float(val):
    """Convert to float."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        val = val.replace('$', '').replace(',', '').strip()
        try:
            return float(val)
        except:
            return 0.0
    return 0.0

def extract_sheet_data(ws, sheet_name):
    """Extract structured data from a worksheet."""
    rows_data = []
    headers = []
    header_row = None

    # Read all rows (up to 500)
    all_rows = []
    for row in ws.iter_rows(max_row=500, values_only=False):
        values = [cell.value for cell in row]
        all_rows.append(values)

    if not all_rows:
        return [], {}

    # Try to find header row - look for rows with keywords
    header_keywords = ['description', 'item', 'category', 'trade', 'cost', 'price',
                       'total', 'labor', 'material', 'quantity', 'qty', 'unit',
                       'amount', 'scope', 'budget', 'allowance', 'estimate']

    for i, row in enumerate(all_rows[:20]):
        str_vals = [str(v).lower() if v else '' for v in row]
        matches = sum(1 for kw in header_keywords if any(kw in sv for sv in str_vals))
        if matches >= 2:
            headers = [str(v).strip() if v else f'col_{j}' for j, v in enumerate(row)]
            header_row = i
            break

    if header_row is None:
        # No clear header - try to extract any rows with text + numbers
        for i, row in enumerate(all_rows):
            non_empty = [v for v in row if v is not None]
            has_text = any(isinstance(v, str) and len(v) > 2 for v in non_empty)
            has_number = any(is_numeric(v) for v in non_empty)
            if has_text and has_number:
                rows_data.append({
                    'row_num': i + 1,
                    'values': [str(v) if v is not None else '' for v in row[:15]]
                })
        return rows_data, {}

    # Extract data rows after header
    line_items = []
    summary = {
        'total_cost': None,
        'labor_total': None,
        'material_total': None,
        'margin_pct': None,
        'overhead': None
    }

    for i in range(header_row + 1, len(all_rows)):
        row = all_rows[i]
        non_empty = [v for v in row if v is not None and str(v).strip()]
        if not non_empty:
            continue

        # Build row dict mapping headers to values
        row_dict = {}
        for j, val in enumerate(row):
            if j < len(headers):
                row_dict[headers[j]] = val
            else:
                row_dict[f'col_{j}'] = val

        # Try to identify this as a line item (has description + cost)
        desc = None
        cost = None
        qty = None
        unit_cost = None
        category = None
        labor = None
        material = None

        for key, val in row_dict.items():
            kl = key.lower()
            if any(k in kl for k in ['description', 'item', 'scope', 'service']):
                if val and isinstance(val, str) and len(val) > 1:
                    desc = str(val).strip()
            elif any(k in kl for k in ['category', 'trade', 'division', 'section']):
                if val and isinstance(val, str):
                    category = str(val).strip()
            elif 'qty' in kl or 'quantity' in kl:
                if is_numeric(val):
                    qty = to_float(val)
            elif 'unit cost' in kl or 'unit price' in kl or 'rate' in kl:
                if is_numeric(val):
                    unit_cost = to_float(val)
            elif 'labor' in kl:
                if is_numeric(val):
                    labor = to_float(val)
            elif 'material' in kl:
                if is_numeric(val):
                    material = to_float(val)
            elif any(k in kl for k in ['total', 'amount', 'cost', 'price', 'budget', 'estimate']):
                if is_numeric(val) and cost is None:
                    cost = to_float(val)

        # Check for summary/total rows
        first_val = str(row[0]).lower() if row[0] else ''
        if any(k in first_val for k in ['total', 'grand total', 'project total', 'subtotal']):
            for val in row:
                if is_numeric(val) and to_float(val) > 0:
                    summary['total_cost'] = to_float(val)
                    break
            continue
        if 'margin' in first_val or 'markup' in first_val or 'profit' in first_val:
            for val in row:
                if is_numeric(val):
                    v = to_float(val)
                    if v < 1:
                        summary['margin_pct'] = v * 100
                    elif v < 100:
                        summary['margin_pct'] = v
                    break
            continue
        if 'overhead' in first_val:
            for val in row:
                if is_numeric(val):
                    summary['overhead'] = to_float(val)
                    break
            continue

        if desc or (cost and cost > 0):
            item = {
                'description': desc or '',
                'category': category,
                'cost': cost,
                'quantity': qty,
                'unit_cost': unit_cost,
                'labor': labor,
                'material': material
            }
            # Clean out None values
            item = {k: v for k, v in item.items() if v is not None}
            if item.get('description') or item.get('cost'):
                line_items.append(item)

    return line_items, summary

def process_file(filepath):
    """Process a single xlsx file and extract data."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        file_data = {
            'sheets': {}
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            items, summary = extract_sheet_data(ws, sheet_name)
            if items:
                file_data['sheets'][sheet_name] = {
                    'line_items': items if isinstance(items, list) and items and isinstance(items[0], dict) and 'description' in items[0] else [],
                    'raw_rows': items if isinstance(items, list) and items and isinstance(items[0], dict) and 'row_num' in items[0] else [],
                    'summary': summary if isinstance(summary, dict) else {}
                }

        wb.close()
        return file_data
    except Exception as e:
        return {'error': str(e)}

def main():
    print("Finding all xlsx files...")
    all_files = find_all_xlsx()
    print(f"Found {len(all_files)} xlsx files")

    # Classify files
    classified = defaultdict(list)
    for f in all_files:
        ftype = classify_file(f)
        classified[ftype].append(f)

    print("\nFile classification:")
    for ftype, files in sorted(classified.items()):
        print(f"  {ftype}: {len(files)}")

    # Process all estimate, allowance, materials, takeoff, and financial files
    priority_types = ['estimate', 'allowance', 'materials_worksheet', 'takeoff', 'financial', 'contracted_work']
    target_files = []
    for ftype in priority_types:
        target_files.extend([(f, ftype) for f in classified.get(ftype, [])])

    # Also process 'other' files that might have pricing data
    for f in classified.get('other', []):
        name = os.path.basename(f).lower()
        if any(k in name for k in ['cost', 'price', 'budget', 'worksheet', 'labor']):
            target_files.append((f, 'other_pricing'))

    print(f"\nProcessing {len(target_files)} priority files...")

    all_estimates = []
    all_line_items = []
    errors = []

    for idx, (filepath, ftype) in enumerate(target_files):
        if idx % 25 == 0:
            print(f"  Processing {idx}/{len(target_files)}...")

        project_name = extract_project_name(filepath)
        project_type = extract_project_type(filepath, project_name)

        result = process_file(filepath)

        if 'error' in result:
            errors.append({'file': filepath, 'error': result['error']})
            continue

        estimate_record = {
            'file': os.path.basename(filepath),
            'file_type': ftype,
            'project_name': project_name,
            'project_type': project_type,
            'path': filepath,
            'sheets': {}
        }

        for sheet_name, sheet_data in result.get('sheets', {}).items():
            items = sheet_data.get('line_items', [])
            summary = sheet_data.get('summary', {})

            estimate_record['sheets'][sheet_name] = {
                'line_item_count': len(items),
                'summary': summary
            }

            for item in items:
                item['source_file'] = os.path.basename(filepath)
                item['project_name'] = project_name
                item['project_type'] = project_type
                item['sheet_name'] = sheet_name
                all_line_items.append(item)

        all_estimates.append(estimate_record)

    print(f"\nExtraction complete:")
    print(f"  Files processed: {len(all_estimates)}")
    print(f"  Total line items: {len(all_line_items)}")
    print(f"  Errors: {len(errors)}")

    # Save extracted estimates
    output = {
        'metadata': {
            'total_xlsx_files': len(all_files),
            'files_processed': len(all_estimates),
            'total_line_items': len(all_line_items),
            'errors': len(errors),
            'file_types_processed': list(set(ft for _, ft in target_files))
        },
        'estimates': all_estimates,
        'line_items': all_line_items,
        'errors': errors[:50]  # Cap errors in output
    }

    with open(os.path.join(OUTPUT_DIR, 'extracted_estimates.json'), 'w') as f:
        json.dump(output, f, indent=2, default=str)
    print(f"Saved extracted_estimates.json")

    # Generate pricing summary
    generate_pricing_summary(all_line_items, all_estimates)

def generate_pricing_summary(line_items, estimates):
    """Generate statistical pricing summary."""

    # Group costs by category/description
    costs_by_description = defaultdict(list)
    costs_by_project_type = defaultdict(list)
    costs_by_category = defaultdict(list)
    labor_rates = []
    material_costs = []
    margins = []

    for item in line_items:
        desc = item.get('description', '').lower().strip()
        cost = item.get('cost')
        project_type = item.get('project_type', 'unknown')
        category = item.get('category', '')

        if cost and cost > 0:
            # Normalize description for grouping
            desc_key = normalize_description(desc)
            if desc_key:
                costs_by_description[desc_key].append({
                    'cost': cost,
                    'project': item.get('project_name', ''),
                    'type': project_type
                })

            costs_by_project_type[project_type].append(cost)

            if category:
                costs_by_category[category.lower().strip()].append(cost)

        if item.get('labor') and item['labor'] > 0:
            labor_rates.append(item['labor'])
        if item.get('material') and item['material'] > 0:
            material_costs.append(item['material'])

    # Extract margins from estimate summaries
    for est in estimates:
        for sheet_name, sheet_data in est.get('sheets', {}).items():
            s = sheet_data.get('summary', {})
            if s.get('margin_pct'):
                margins.append(s['margin_pct'])

    # Build summary
    pricing_by_item = {}
    for desc, entries in costs_by_description.items():
        costs = [e['cost'] for e in entries]
        if len(costs) >= 1 and costs[0] > 0:
            pricing_by_item[desc] = {
                'count': len(costs),
                'min': round(min(costs), 2),
                'max': round(max(costs), 2),
                'avg': round(mean(costs), 2),
                'median': round(median(costs), 2) if len(costs) > 1 else round(costs[0], 2),
                'project_types': list(set(e['type'] for e in entries))
            }

    # Sort by frequency
    pricing_by_item = dict(sorted(pricing_by_item.items(), key=lambda x: x[1]['count'], reverse=True))

    pricing_by_project = {}
    for ptype, costs in costs_by_project_type.items():
        if costs:
            pricing_by_project[ptype] = {
                'project_count': len(set(costs)),
                'total_line_items': len(costs),
                'avg_line_item_cost': round(mean(costs), 2),
                'total_cost_range': {
                    'min': round(min(costs), 2),
                    'max': round(max(costs), 2)
                }
            }

    summary = {
        'metadata': {
            'total_line_items_with_cost': sum(len(v) for v in costs_by_description.values()),
            'unique_item_descriptions': len(pricing_by_item),
            'project_types_found': len(pricing_by_project),
            'labor_entries': len(labor_rates),
            'material_entries': len(material_costs),
            'margin_entries': len(margins)
        },
        'pricing_by_item_type': pricing_by_item,
        'pricing_by_project_type': pricing_by_project,
        'pricing_by_category': {
            cat: {
                'count': len(costs),
                'avg': round(mean(costs), 2),
                'min': round(min(costs), 2),
                'max': round(max(costs), 2)
            }
            for cat, costs in sorted(costs_by_category.items(), key=lambda x: len(x[1]), reverse=True)
            if costs
        },
        'labor_analysis': {
            'entries': len(labor_rates),
            'avg': round(mean(labor_rates), 2) if labor_rates else 0,
            'min': round(min(labor_rates), 2) if labor_rates else 0,
            'max': round(max(labor_rates), 2) if labor_rates else 0,
            'median': round(median(labor_rates), 2) if len(labor_rates) > 1 else (labor_rates[0] if labor_rates else 0)
        },
        'material_analysis': {
            'entries': len(material_costs),
            'avg': round(mean(material_costs), 2) if material_costs else 0,
            'min': round(min(material_costs), 2) if material_costs else 0,
            'max': round(max(material_costs), 2) if material_costs else 0,
            'median': round(median(material_costs), 2) if len(material_costs) > 1 else (material_costs[0] if material_costs else 0)
        },
        'margin_analysis': {
            'entries': len(margins),
            'avg_pct': round(mean(margins), 2) if margins else 0,
            'min_pct': round(min(margins), 2) if margins else 0,
            'max_pct': round(max(margins), 2) if margins else 0
        }
    }

    with open(os.path.join(OUTPUT_DIR, 'pricing_summary.json'), 'w') as f:
        json.dump(summary, f, indent=2)
    print(f"Saved pricing_summary.json")

def normalize_description(desc):
    """Normalize line item descriptions for grouping."""
    if not desc or len(desc) < 3:
        return None

    # Remove numbers, dates, client names
    desc = re.sub(r'\d{1,2}/\d{1,2}/\d{2,4}', '', desc)
    desc = re.sub(r'\$[\d,.]+', '', desc)

    # Common normalizations
    mappings = {
        'framing': ['framing', 'frame', 'structural framing'],
        'roofing': ['roofing', 'roof', 'shingles', 'metal roof'],
        'electrical': ['electrical', 'electric', 'wiring', 'outlets'],
        'plumbing': ['plumbing', 'plumb', 'pipes', 'fixtures'],
        'hvac': ['hvac', 'heating', 'cooling', 'a/c', 'air conditioning', 'ductwork'],
        'insulation': ['insulation', 'insulate'],
        'drywall': ['drywall', 'sheetrock', 'sheet rock'],
        'painting': ['painting', 'paint', 'primer', 'stain'],
        'flooring': ['flooring', 'floor', 'hardwood', 'lvp', 'tile floor', 'carpet'],
        'concrete': ['concrete', 'foundation', 'slab', 'footing'],
        'cabinets': ['cabinet', 'cabinetry'],
        'countertops': ['countertop', 'granite', 'quartz', 'counter top'],
        'demolition': ['demolition', 'demo', 'tear out', 'removal'],
        'trim_finish': ['trim', 'finish carpentry', 'molding', 'baseboard', 'crown'],
        'doors': ['door', 'doors'],
        'windows': ['window', 'windows'],
        'decking': ['decking', 'deck boards', 'trex', 'composite deck'],
        'railing': ['railing', 'rail', 'handrail', 'baluster'],
        'gutters': ['gutter', 'gutters', 'downspout'],
        'tile': ['tile', 'tiling', 'backsplash', 'shower tile'],
        'permits': ['permit', 'permits', 'permitting'],
        'cleanup': ['cleanup', 'clean up', 'debris', 'dumpster', 'hauling'],
        'grading': ['grading', 'excavation', 'site prep', 'earthwork'],
        'siding': ['siding', 'hardie', 'board siding'],
        'ceiling': ['ceiling', 'bead board', 'tongue and groove'],
        'appliances': ['appliance', 'appliances'],
        'lighting': ['lighting', 'light fixture', 'fixtures'],
        'landscaping': ['landscaping', 'landscape', 'sod', 'turf'],
        'masonry': ['masonry', 'brick', 'stone', 'block'],
        'waterproofing': ['waterproof', 'moisture barrier'],
        'stairs_steps': ['stairs', 'steps', 'staircase'],
        'fencing': ['fence', 'fencing'],
        'garage_door': ['garage door'],
        'septic': ['septic'],
        'well': ['well'],
    }

    desc_lower = desc.lower()
    for normalized, keywords in mappings.items():
        for kw in keywords:
            if kw in desc_lower:
                return normalized

    # If no mapping found, return cleaned desc (truncate for grouping)
    clean = re.sub(r'[^a-z\s]', '', desc_lower).strip()
    words = clean.split()[:4]
    return ' '.join(words) if words else None

if __name__ == '__main__':
    main()
